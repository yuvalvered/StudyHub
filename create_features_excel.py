import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Colors
HEADER_BG = "2C3E50"
HEADER_FG = "FFFFFF"
SECTION_BG = "3498DB"
SECTION_FG = "FFFFFF"
DONE_BG = "D5F5E3"
PARTIAL_BG = "FEF9E7"
NOT_IMPL_BG = "FADBD8"
ALT_ROW = "F2F3F4"
WHITE = "FFFFFF"

def header_style(cell, bg=HEADER_BG, fg=HEADER_FG, size=12):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fg, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def section_style(cell, bg=SECTION_BG):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=SECTION_FG, size=11, name="Calibri")
    cell.alignment = Alignment(horizontal="right", vertical="center")

def data_style(cell, bg=WHITE, bold=False):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, size=10, name="Calibri")
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

def border_all(cell):
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def status_bg(status):
    if "✅" in status:
        return DONE_BG
    elif "⚠️" in status:
        return PARTIAL_BG
    else:
        return NOT_IMPL_BG

# ──────────────────────────────────────────────
# Sheet 1: Frontend Features
# ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "פיצ'רים - Frontend"

headers = ["#", "פיצ'ר", "תיאור", "מיקום בקוד", "סטטוס"]
col_widths = [5, 28, 50, 35, 12]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=1, column=col, value=h)
    header_style(cell)
    border_all(cell)
    ws1.column_dimensions[get_column_letter(col)].width = w

ws1.row_dimensions[1].height = 30

frontend_sections = [
    ("אימות וניהול חשבון", [
        (1, "הרשמה", "רישום עם אימייל BGU, אימות אימייל, ולידציה של סיסמה", "/register", "✅ מלא"),
        (2, "התחברות", "אימות עם שם משתמש/סיסמה, קבלת JWT token", "/login", "✅ מלא"),
        (3, "אימות אימייל", "שליחת קישור אימות בהרשמה, אימות token", "/verify-email", "✅ מלא"),
        (4, "שכחתי סיסמה", "בקשת איפוס סיסמה דרך אימייל", "/forgot-password", "✅ מלא"),
        (5, "איפוס סיסמה", "קביעת סיסמה חדשה דרך token באימייל", "/reset-password", "✅ מלא"),
        (6, "ניהול פרופיל", "עריכת שם, מחלקה, שנה, ביוגרפיה", "/profile", "✅ מלא"),
        (7, "העלאת תמונת פרופיל", "העלאת תמונה (JPG/PNG/GIF עד 5MB)", "/profile", "✅ מלא"),
        (8, "סטטיסטיקות משתמש", "הצגת מספר העלאות, הורדות, ציון ממוצע", "/profile", "✅ מלא"),
        (9, "הגדרות התראות", "הגדרת אילו התראות לקבל (אפליקציה / אימייל)", "/profile", "✅ מלא"),
    ]),
    ("דשבורד וניהול קורסים", [
        (10, "דשבורד ראשי", "הצגת קורסים פעילים עם כרטיסיות וסטטיסטיקות", "/dashboard", "✅ מלא"),
        (11, "הוספת קורס", "חיפוש והרשמה לקורסים מהמאגר", "/dashboard", "✅ מלא"),
        (12, "הסרת קורס", "הסרת קורס מהרשימה האישית", "/dashboard", "✅ מלא"),
        (13, "חיפוש קורסים", "חיפוש גלובלי למציאת קורסים", "/dashboard", "✅ מלא"),
        (14, "סטטיסטיקות דשבורד", "קורסים פעילים, העלאות, הורדות, שותפי לימוד", "/dashboard", "✅ מלא"),
        (15, "פאנל אדמין", "הוספה/מחיקה של קורסים (אדמין בלבד)", "/dashboard sidebar", "✅ מלא"),
        (16, "רשימת קורסים", "צפייה בכל הקורסים הזמינים", "/courses", "⚠️ חלקי"),
    ]),
    ("חומרי לימוד", [
        (17, "קטגוריות חומרים", "7 סוגים: סיכומים, הרצאות, תרגילים, שיעורי בית, בחינות, בחנים, QuizMe", "/courses/[id]", "✅ מלא"),
        (18, "רשימת חומרים", "עיון בחומרים לפי קטגוריה עם אפשרויות מיון", "/courses/[id]/materials/[type]", "✅ מלא"),
        (19, "העלאת חומרים", "העלאת קבצים (PDF, DOCX, PPTX, XLSX) עם כותרת ותיאור", "/courses/[id]/materials/[type]", "✅ מלא"),
        (20, "תצוגה מקדימה", "תצוגת PDF וקבצי Office (דרך המרה ל-PDF)", "/courses/[id]/materials/[type]/[id]", "✅ מלא"),
        (21, "הורדת חומר", "הורדת הקובץ המקורי (כולל תמיכה בשמות עבריים)", "/courses/[id]/materials/[type]/[id]", "✅ מלא"),
        (22, "מחיקת חומר", "מחיקת חומר (בעלים בלבד)", "/courses/[id]/materials/[type]/[id]", "✅ מלא"),
        (23, "דירוג חומרים", "דירוג 1-5 כוכבים עם תגובת ביקורת אופציונלית", "/courses/[id]/materials/[type]/[id]", "✅ מלא"),
        (24, "דיווח על חומר", "דיווח על תוכן לא הולם", "/courses/[id]/materials/[type]/[id]", "✅ מלא"),
    ]),
    ("חיפוש", [
        (25, "חיפוש טקסט מלא", "חיפוש חומרים לפי כותרת, תיאור ותוכן הקובץ", "/courses/[id] header", "✅ מלא"),
        (26, "תוצאות חיפוש", "הצגת תוצאות עם קטעים מודגשים ואפשרויות מיון", "/courses/[id]", "✅ מלא"),
        (27, "פילטרים בחיפוש", "סינון לפי קורס, סוג חומר, מיון לפי רלוונטיות/תאריך/דירוג", "/courses/[id]", "✅ מלא"),
    ]),
    ("דיונים וקהילה", [
        (28, "יצירת דיון", "פתיחת דיון חדש ברמת הקורס", "/courses/[id]", "✅ מלא"),
        (29, "צפייה בדיונים", "רשימת כל הדיונים בקורס", "/courses/[id]", "✅ מלא"),
        (30, "תגובות בדיון", "הוספת תגובות לדיון", "/courses/[id] modal", "✅ מלא"),
        (31, "תגובות מקוננות", "תשובה לתגובה ספציפית (threads)", "/courses/[id] modal", "✅ מלא"),
        (32, "הצבעה על תגובות", "לייק/דיסלייק על תגובות", "/courses/[id] modal", "✅ מלא"),
        (33, "מיון תגובות", "מיון לפי חדש ביותר או הכי מדורג", "/courses/[id] modal", "✅ מלא"),
    ]),
    ("שותפי לימוד", [
        (34, "מצב שותף לימוד", "סימון 'מחפש שותף לימוד' בקורסים ספציפיים", "/dashboard", "✅ מלא"),
        (35, "מציאת שותפים", "הצגת משתמשים שמחפשים שותף לימוד בקורס", "/courses/[id]", "⚠️ חלקי"),
    ]),
    ("AI", [
        (36, "שאל את הAI", "שאלות על חומר ספציפי או כל חומרי הקורס", "/courses/[id]/materials/[type]/[id]", "✅ מלא"),
        (37, "בחירת מודל AI", "בחירה בין Gemini או Ollama", "/courses/[id]/materials/[type]/[id]", "✅ מלא"),
        (38, "ממשק צ'אט", "שיחת AI אינטראקטיבית עם שאלות המשך", "/courses/[id]/materials/[type]/[id]", "✅ מלא"),
    ]),
    ("התראות", [
        (39, "פעמון התראות", "הצגת מספר ההתראות שלא נקראו", "Global Header", "✅ מלא"),
        (40, "התראה על תגובה לחומר", "התראה כשמישהו מגיב על החומר שלך", "אוטומטי", "✅ מלא"),
        (41, "התראה על דירוג", "התראה כשמישהו מדרג את החומר שלך", "אוטומטי", "✅ מלא"),
        (42, "התראה על תשובה לתגובה", "התראה כשמישהו עונה לתגובה שלך", "אוטומטי", "✅ מלא"),
    ]),
]

row = 2
for section_name, features in frontend_sections:
    # Section header
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws1.cell(row=row, column=1, value=section_name)
    section_style(cell)
    border_all(cell)
    ws1.row_dimensions[row].height = 22
    row += 1

    for i, (num, name, desc, location, status) in enumerate(features):
        bg = DONE_BG if "✅" in status else (PARTIAL_BG if "⚠️" in status else NOT_IMPL_BG)
        row_bg = bg if i % 2 == 0 else (ALT_ROW if "✅" in status else bg)
        for col, val in enumerate([num, name, desc, location, status], 1):
            cell = ws1.cell(row=row, column=col, value=val)
            data_style(cell, bg=bg)
            border_all(cell)
        ws1.row_dimensions[row].height = 20
        row += 1

ws1.freeze_panes = "A2"


# ──────────────────────────────────────────────
# Sheet 2: Backend API
# ──────────────────────────────────────────────
ws2 = wb.create_sheet("פיצ'רים - Backend API")

headers2 = ["#", "Method", "Endpoint", "תיאור", "קובץ", "סטטוס"]
col_widths2 = [5, 10, 40, 45, 20, 12]

for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=col, value=h)
    header_style(cell)
    border_all(cell)
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.row_dimensions[1].height = 30

METHOD_COLORS = {
    "GET": "27AE60", "POST": "2980B9", "PUT": "E67E22",
    "DELETE": "E74C3C", "PATCH": "8E44AD"
}

backend_sections = [
    ("אימות ואבטחה", [
        (1, "POST", "/auth/register", "יצירת חשבון משתמש חדש", "auth.py", "✅ מלא"),
        (2, "POST", "/auth/login", "אימות וקבלת JWT tokens", "auth.py", "✅ מלא"),
        (3, "POST", "/auth/login/json", "התחברות דרך JSON", "auth.py", "✅ מלא"),
        (4, "POST", "/auth/forgot-password", "בקשת איפוס סיסמה דרך אימייל", "auth.py", "✅ מלא"),
        (5, "POST", "/auth/password-reset", "אישור איפוס סיסמה עם token", "auth.py", "✅ מלא"),
        (6, "POST", "/auth/verify-email", "אימות כתובת אימייל עם token", "auth.py", "✅ מלא"),
        (7, "POST", "/auth/refresh", "רענון JWT token", "auth.py", "⚠️ לא ממומש"),
    ]),
    ("ניהול קורסים", [
        (8, "POST", "/courses", "יצירת קורס חדש (אדמין)", "courses.py", "✅ מלא"),
        (9, "GET", "/courses", "רשימת קורסים עם פילטרים ועימוד", "courses.py", "✅ מלא"),
        (10, "GET", "/courses/{id}", "פרטי קורס ספציפי", "courses.py", "✅ מלא"),
        (11, "PUT", "/courses/{id}", "עדכון פרטי קורס", "courses.py", "✅ מלא"),
        (12, "DELETE", "/courses/{id}", "מחיקת קורס וחומריו", "courses.py", "✅ מלא"),
        (13, "POST", "/courses/{id}/enroll", "הרשמה לקורס", "courses.py", "✅ מלא"),
        (14, "DELETE", "/courses/{id}/enroll", "ביטול הרשמה לקורס", "courses.py", "✅ מלא"),
        (15, "GET", "/courses/{id}/materials", "כל החומרים של קורס", "courses.py", "✅ מלא"),
    ]),
    ("חומרי לימוד", [
        (16, "GET", "/materials", "רשימת חומרים עם פילטרים ועימוד", "materials.py", "✅ מלא"),
        (17, "POST", "/courses/{id}/materials/upload", "העלאת חומר לימוד עם קובץ", "materials.py", "✅ מלא"),
        (18, "GET", "/materials/{id}", "פרטי חומר עם מידע על המעלה", "materials.py", "✅ מלא"),
        (19, "PUT", "/materials/{id}", "עדכון מטא-דאטה של חומר", "materials.py", "✅ מלא"),
        (20, "DELETE", "/materials/{id}", "מחיקת חומר (בעלים/אדמין בלבד)", "materials.py", "✅ מלא"),
        (21, "GET", "/materials/{id}/preview", "תצוגה מקדימה של קובץ (PDF / המרה)", "materials.py", "✅ מלא"),
        (22, "GET", "/materials/{id}/download", "הורדת קובץ החומר", "materials.py", "✅ מלא"),
    ]),
    ("דירוגים", [
        (23, "POST", "/materials/{id}/rate", "יצירת דירוג לחומר (1-5 כוכבים)", "ratings.py", "✅ מלא"),
        (24, "PUT", "/materials/{id}/rate", "עדכון דירוג קיים", "ratings.py", "✅ מלא"),
        (25, "DELETE", "/materials/{id}/rate", "מחיקת דירוג", "ratings.py", "✅ מלא"),
        (26, "GET", "/materials/{id}/ratings", "כל הדירוגים של חומר", "ratings.py", "✅ מלא"),
    ]),
    ("דיונים ותגובות", [
        (27, "POST", "/courses/{id}/discussions", "יצירת דיון בקורס", "discussions.py", "✅ מלא"),
        (28, "GET", "/courses/{id}/discussions", "רשימת דיונים בקורס", "discussions.py", "✅ מלא"),
        (29, "GET", "/discussions/{id}", "פרטי דיון", "discussions.py", "✅ מלא"),
        (30, "PUT", "/discussions/{id}", "עדכון דיון", "discussions.py", "✅ מלא"),
        (31, "DELETE", "/discussions/{id}", "מחיקת דיון", "discussions.py", "✅ מלא"),
        (32, "POST", "/discussions/{id}/comments", "הוספת תגובה לדיון", "comments.py", "✅ מלא"),
        (33, "GET", "/discussions/{id}/comments", "רשימת תגובות עם threading", "comments.py", "✅ מלא"),
        (34, "PUT", "/comments/{id}", "עדכון תגובה", "comments.py", "✅ מלא"),
        (35, "DELETE", "/comments/{id}", "מחיקת תגובה", "comments.py", "✅ מלא"),
        (36, "POST", "/comments/{id}/vote", "הצבעה על תגובה (לייק/דיסלייק)", "comments.py", "✅ מלא"),
    ]),
    ("ניהול משתמשים", [
        (37, "GET", "/users/me", "פרטי פרופיל משתמש נוכחי", "users.py", "✅ מלא"),
        (38, "GET", "/users/me/stats", "סטטיסטיקות משתמש", "users.py", "✅ מלא"),
        (39, "PUT", "/users/me", "עדכון פרופיל משתמש", "users.py", "✅ מלא"),
        (40, "POST", "/users/me/profile-image", "העלאת תמונת פרופיל", "users.py", "✅ מלא"),
        (41, "GET", "/users/me/courses", "קורסים שנרשמתי אליהם", "users.py", "✅ מלא"),
        (42, "PUT", "/users/{id}/courses/{course_id}", "עדכון הרשמה (שותף לימוד)", "users.py", "✅ מלא"),
    ]),
    ("חיפוש", [
        (43, "GET", "/search/materials", "חיפוש טקסט מלא עם פילטרים ומיון", "search.py", "✅ מלא"),
    ]),
    ("התראות", [
        (44, "GET", "/notifications", "התראות משתמש עם עימוד", "notifications.py", "✅ מלא"),
        (45, "GET", "/notifications/unread", "התראות שלא נקראו בלבד", "notifications.py", "✅ מלא"),
        (46, "PUT", "/notifications/{id}", "סימון התראה כנקראה", "notifications.py", "✅ מלא"),
        (47, "DELETE", "/notifications/{id}", "מחיקת התראה", "notifications.py", "✅ מלא"),
        (48, "GET", "/notifications/settings", "קבלת העדפות התראות", "notifications.py", "✅ מלא"),
        (49, "PUT", "/notifications/settings", "עדכון העדפות התראות", "notifications.py", "✅ מלא"),
    ]),
    ("AI", [
        (50, "POST", "/ai/ask", "שאלה על חומר/קורס (Gemini או Ollama)", "ai.py", "✅ מלא"),
        (51, "POST", "/ai/summarize", "סיכום תוכן מסמך", "ai.py", "✅ מלא"),
        (52, "POST", "/ai/generate-questions", "יצירת שאלות אימון אוטומטית", "ai.py", "✅ מלא"),
    ]),
]

row = 2
for section_name, endpoints in backend_sections:
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws2.cell(row=row, column=1, value=section_name)
    section_style(cell)
    border_all(cell)
    ws2.row_dimensions[row].height = 22
    row += 1

    for num, method, endpoint, desc, file, status in endpoints:
        bg = DONE_BG if "✅" in status else PARTIAL_BG
        for col, val in enumerate([num, method, endpoint, desc, file, status], 1):
            cell = ws2.cell(row=row, column=col, value=val)
            data_style(cell, bg=bg)
            if col == 2:  # Method
                method_color = METHOD_COLORS.get(method, "7F8C8D")
                cell.fill = PatternFill("solid", fgColor=method_color)
                cell.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            border_all(cell)
        ws2.row_dimensions[row].height = 20
        row += 1

ws2.freeze_panes = "A2"


# ──────────────────────────────────────────────
# Sheet 3: Services
# ──────────────────────────────────────────────
ws3 = wb.create_sheet("שירותים ומודלים")

headers3 = ["רכיב", "שם", "תיאור", "פונקציות מרכזיות", "סטטוס"]
col_widths3 = [18, 30, 45, 50, 12]

for col, (h, w) in enumerate(zip(headers3, col_widths3), 1):
    cell = ws3.cell(row=1, column=col, value=h)
    header_style(cell)
    border_all(cell)
    ws3.column_dimensions[get_column_letter(col)].width = w

ws3.row_dimensions[1].height = 30

services_data = [
    ("Services", [
        ("שירות אימות", "auth_service.py", "הרשמה, התחברות, איפוס סיסמה, JWT tokens", "register_user(), login_user(), reset_password(), verify_email()", "✅ מלא"),
        ("שירות קורסים", "course_service.py", "CRUD קורסים וניהול הרשמות", "create_course(), get_courses(), enroll_user()", "✅ מלא"),
        ("שירות חומרים", "material_service.py", "העלאה, הורדה, תצוגה מקדימה, מטא-דאטה", "create_material(), get_material(), increment_download_count()", "✅ מלא"),
        ("שירות קבצים", "file_service.py", "שמירת קבצים, חילוץ טקסט, ספירת עמודים", "extract_file_text(), extract_pdf_text(), fix_hebrew_direction()", "✅ מלא"),
        ("שירות חיפוש", "search_service.py", "Full-Text Search, ILIKE, Trigram, תמיכה בעברית", "search_materials(), generate_snippet(), handle_hebrew_variants()", "✅ מלא"),
        ("שירות התראות", "notification_service.py", "ניהול התראות, שליחת אימיילים, הגדרות", "create_notification(), get_unread(), update_settings()", "✅ מלא"),
        ("שירות אימייל", "email_service.py", "שליחת אימיילים לאימות, איפוס, התראות", "send_verification_email(), send_password_reset(), send_notification()", "✅ מלא"),
        ("שירות AI Q&A", "document_qa_service.py", "מענה על שאלות דרך Gemini או Ollama", "answer_question(), answer_question_multi_docs(), is_model_available()", "✅ מלא"),
        ("המרת PDF", "pdf_conversion_service.py", "המרת DOCX/PPTX ל-PDF לתצוגה מקדימה (LibreOffice)", "convert_to_pdf(), is_available(), can_convert()", "✅ מלא"),
        ("חילוץ נושאים", "topic_extraction_service.py", "חילוץ נושאים/מילות מפתח ממסמכים (Ollama)", "extract_topics()", "✅ מלא"),
    ]),
    ("Database Models", [
        ("מודל משתמש", "User", "חשבון משתמש עם פרופיל וסטטיסטיקות", "id, username, email, full_name, department, year, bio, profile_image", "✅ מלא"),
        ("מודל קורס", "Course", "קורסים אקדמיים עם מטא-דאטה", "id, name, department, course_number, description", "✅ מלא"),
        ("מודל חומר", "Material", "חומרי לימוד ומסמכים", "id, title, file_path, file_content_text, material_type, avg_rating", "✅ מלא"),
        ("מודל דירוג", "Rating", "דירוגי משתמשים לחומרים", "id, material_id, user_id, score, comment", "✅ מלא"),
        ("מודל דיון", "Discussion", "דיוני קורס", "id, course_id, user_id, title, content", "✅ מלא"),
        ("מודל תגובה", "Comment", "תגובות עם תמיכה ב-threading", "id, discussion_id, user_id, parent_id, content, vote_count", "✅ מלא"),
        ("מודל הצבעה", "CommentVote", "הצבעות על תגובות", "id, comment_id, user_id, vote_type", "✅ מלא"),
        ("מודל דיווח", "MaterialReport", "דיווחים על תוכן לא הולם", "id, material_id, user_id, reason", "✅ מלא"),
        ("מודל הרשמה", "UserCourse", "טבלת חיבור משתמש-קורס עם דגל שותף לימוד", "user_id, course_id, is_looking_for_partner", "✅ מלא"),
        ("מודל התראה", "Notification", "התראות משתמש", "id, user_id, type, content, is_read", "✅ מלא"),
        ("הגדרות התראות", "NotificationSettings", "העדפות התראות לכל משתמש", "user_id, comment_in_app, rating_in_app, reply_email...", "✅ מלא"),
    ]),
]

row = 2
for section_name, items in services_data:
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws3.cell(row=row, column=1, value=section_name)
    section_style(cell)
    border_all(cell)
    ws3.row_dimensions[row].height = 22
    row += 1

    for comp, name, desc, funcs, status in items:
        bg = DONE_BG if "✅" in status else PARTIAL_BG
        for col, val in enumerate([comp, name, desc, funcs, status], 1):
            cell = ws3.cell(row=row, column=col, value=val)
            data_style(cell, bg=bg)
            border_all(cell)
        ws3.row_dimensions[row].height = 22
        row += 1

ws3.freeze_panes = "A2"


# ──────────────────────────────────────────────
# Sheet 4: Summary
# ──────────────────────────────────────────────
ws4 = wb.create_sheet("סיכום")

summary_data = [
    ("קטגוריה", "סה\"כ פיצ'רים", "מלא ✅", "חלקי ⚠️", "לא ממומש ❌"),
    ("Frontend", 42, 39, 3, 0),
    ("Backend API", 52, 51, 1, 0),
    ("Services", 10, 10, 0, 0),
    ("Database Models", 11, 11, 0, 0),
    ("סה\"כ", 115, 111, 4, 0),
]

col_widths4 = [25, 20, 15, 15, 18]
for col, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(col)].width = w

for row_idx, row_data in enumerate(summary_data, 1):
    for col, val in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col, value=val)
        if row_idx == 1:
            header_style(cell)
        elif row_idx == len(summary_data):
            cell.fill = PatternFill("solid", fgColor="2C3E50")
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.fill = PatternFill("solid", fgColor=DONE_BG if col > 1 else WHITE)
            cell.font = Font(size=11, name="Calibri")
            cell.alignment = Alignment(horizontal="center" if col > 1 else "right", vertical="center")
        border_all(cell)
    ws4.row_dimensions[row_idx].height = 25

ws4.title = "סיכום"

output_path = r"c:\Users\yuval\Desktop\studies\forth year\final project\StudyHub\StudyHub_Features.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
